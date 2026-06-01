import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
import openpyxl
import numpy as np
import jmespath

wbname = 'budgetmanager.xlsx'


# ==========================================
# BACKEND (Zachowany zgodnie z wytycznymi)
# ==========================================

class Wydatek:
    def __init__(self, data, kategoria, kwota, opis):
        self.data = data
        self.kategoria = kategoria
        self.kwota = kwota
        self.opis = opis
        pass


class BudgetManager:
    def __init__(self):
        self.wydatki = []
        try:
            wb = openpyxl.load_workbook(wbname)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                data = row[0].value
                kategoria = row[1].value
                kwota = row[2].value
                opis = row[3].value
                if data is not None:  # zabezpieczenie przed pustymi wierszami
                    self.wydatki.append(Wydatek(data, kategoria, kwota, opis))
            wb.save(wbname)
        except FileNotFoundError:
            # Tworzy nowy plik, jeśli nie istnieje
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Data", "Kategoria", "Kwota", "Opis"])
            wb.save(wbname)
        pass

    def dodaj_wydatek(self, wydatek_obj):
        # Poprawiona wielkość litery zmiennej z przekazanego parametru
        data = wydatek_obj.data
        kategoria = wydatek_obj.kategoria
        kwota = wydatek_obj.kwota
        opis = wydatek_obj.opis
        self.wydatki.append(Wydatek(data, kategoria, kwota, opis))

        # Zapis do pliku excel przy dodawaniu
        try:
            wb = openpyxl.load_workbook(wbname)
            ws = wb.active
            ws.append([data, kategoria, kwota, opis])
            wb.save(wbname)
        except Exception:
            pass
        pass

    def oblicz_sume_kategorii(self, kategoriaSzukana):
        suma = 0
        for konkretnywydatek in self.wydatki:
            if konkretnywydatek.kategoria == kategoriaSzukana:
                try:
                    suma += int(konkretnywydatek.kwota)
                except (ValueError, TypeError):
                    try:
                        suma += float(konkretnywydatek.kwota)
                    except:
                        pass
        return suma

    def wczytaj_z_excela(self, nazwaexcela):
        self.wydatki = []
        wb = openpyxl.load_workbook(nazwaexcela)
        ws = wb.active
        # Poprawiono min_row=2, aby nie wczytywać nagłówków jako danych
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            data = row[0].value
            kategoria = row[1].value
            kwota = row[2].value
            opis = row[3].value
            if data is not None:
                self.wydatki.append(Wydatek(data, kategoria, kwota, opis))
        wb.save(nazwaexcela)
        pass

    def czysc_dane(self):
        self.wydatki = []
        # Wyczyszczenie pliku Excel (zostają tylko nagłówki)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Data", "Kategoria", "Kwota", "Opis"])
        wb.save(wbname)
        pass


# ==========================================
# FRONTEND + INTEGRACJA
# ==========================================

class AddExpenseDialog(QDialog):
    """Okno dialogowe do dodawania nowego wydatku"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Dodaj nowy wydatek")
        self.resize(300, 250)
        self.setStyleSheet("background-color: #313244; color: white; font-size: 14px;")

        layout = QVBoxLayout(self)

        self.data_input = QLineEdit(self)
        self.data_input.setPlaceholderText("Data (np. 2026-06-01)")

        self.kat_input = QLineEdit(self)
        self.kat_input.setPlaceholderText("Kategoria (np. Jedzenie)")

        self.kwota_input = QLineEdit(self)
        self.kwota_input.setPlaceholderText("Kwota (np. 50)")

        self.opis_input = QLineEdit(self)
        self.opis_input.setPlaceholderText("Opis (np. Zakupy w markecie)")

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)

        layout.addWidget(QLabel("Data:"))
        layout.addWidget(self.data_input)
        layout.addWidget(QLabel("Kategoria:"))
        layout.addWidget(self.kat_input)
        layout.addWidget(QLabel("Kwota:"))
        layout.addWidget(self.kwota_input)
        layout.addWidget(QLabel("Opis:"))
        layout.addWidget(self.opis_input)
        layout.addWidget(btn_box)

    def get_data(self):
        return (
            self.data_input.text(),
            self.kat_input.text(),
            self.kwota_input.text(),
            self.opis_input.text()
        )


class BudgetGUI(QMainWindow):
    def __init__(self):
        super().__init__()

        # Inicjalizacja managera danych
        self.manager = BudgetManager()

        self.setWindowTitle("💰 Budget Manager")
        self.resize(1200, 700)

        self.setStyleSheet("""
            QMainWindow {
                background-color: #1e1e2e;
            }

            QLabel {
                color: white;
            }

            QPushButton {
                background-color: #89b4fa;
                color: black;
                border-radius: 10px;
                padding: 10px;
                font-weight: bold;
            }

            QPushButton:hover {
                background-color: #74c7ec;
            }

            QTableWidget {
                background-color: #313244;
                color: white;
                gridline-color: #45475a;
                border-radius: 10px;
            }

            QHeaderView::section {
                background-color: #45475a;
                color: white;
                padding: 5px;
                border: none;
            }

            QListWidget {
                background-color: #181825;
                color: white;
                border: none;
                font-size: 15px;
            }
        """)

        central = QWidget()
        self.setCentralWidget(central)

        layout = QHBoxLayout(central)

        # MENU LEWE
        menu = QListWidget()
        menu.addItems([
            "📊 Dashboard",
            "💸 Wydatki",
            "📁 Kategorie",
            "📈 Statystyki"
        ])
        menu.setMaximumWidth(220)

        layout.addWidget(menu)

        # PRAWA CZĘŚĆ
        right = QVBoxLayout()

        title = QLabel("💰 Budget Manager")
        title.setFont(QFont("Segoe UI", 24))

        right.addWidget(title)

        # Karty statystyk
        cards = QHBoxLayout()

        self.card_labels = []
        for i in range(3):
            card = QFrame()
            card.setStyleSheet("""
                background:#313244;
                border-radius:15px;
                color:white;
            """)
            card_layout = QVBoxLayout(card)
            label = QLabel()
            label.setAlignment(Qt.AlignCenter)
            label.setFont(QFont("Segoe UI", 14))
            card_layout.addWidget(label)
            cards.addWidget(card)
            self.card_labels.append(label)

        right.addLayout(cards)

        # tabela
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(
            ["Data", "Kategoria", "Kwota", "Opis"]
        )
        self.table.horizontalHeader().setStretchLastSection(True)
        right.addWidget(self.table)

        # przyciski
        buttons = QHBoxLayout()

        add_btn = QPushButton("➕ Dodaj wydatek")
        load_btn = QPushButton("📂 Wczytaj Excel")
        clear_btn = QPushButton("🗑 Wyczyść")

        buttons.addWidget(add_btn)
        buttons.addWidget(load_btn)
        buttons.addWidget(clear_btn)

        right.addLayout(buttons)
        layout.addLayout(right)

        # Podpięcie akcji pod przyciski
        add_btn.clicked.connect(self.action_dodaj_wydatek)
        load_btn.clicked.connect(self.action_wczytaj_excel)
        clear_btn.clicked.connect(self.action_wyczysc)

        # Odświeżenie widoku na start programu
        self.odswiez_widok()

    def odswiez_widok(self):
        """Aktualizuje tabelę oraz kafelki podsumowujące na podstawie danych z managera"""
        # 1. Aktualizacja Tabeli
        self.table.setRowCount(0)
        calkowita_suma = 0

        for w in self.manager.wydatki:
            row_idx = self.table.rowCount()
            self.table.insertRow(row_idx)
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(w.data)))
            self.table.setItem(row_idx, 1, QTableWidgetItem(str(w.kategoria)))
            self.table.setItem(row_idx, 2, QTableWidgetItem(str(w.kwota)))
            self.table.setItem(row_idx, 3, QTableWidgetItem(str(w.opis)))

            try:
                calkowita_suma += float(w.kwota)
            except:
                pass

        # 2. Aktualizacja Kart Statystyk (korzystając z metod klasy BudgetManager)
        suma_jedzenie = self.manager.oblicz_sume_kategorii("Jedzenie")
        suma_transport = self.manager.oblicz_sume_kategorii("Transport")

        self.card_labels[0].setText(f"Suma\n{calkowita_suma:.2f} zł")
        self.card_labels[1].setText(f"Jedzenie\n{suma_jedzenie:.2f} zł")
        self.card_labels[2].setText(f"Transport\n{suma_transport:.2f} zł")

    def action_dodaj_wydatek(self):
        dialog = AddExpenseDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data, kategoria, kwota, opis = dialog.get_data()
            if data and kategoria and kwota:  # Prosta walidacja czy pola nie są puste
                nowy = Wydatek(data, kategoria, kwota, opis)
                self.manager.dodaj_wydatek(nowy)
                self.odswiez_widok()

    def action_wczytaj_excel(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Wybierz plik Excel", "", "Excel Files (*.xlsx)")
        if file_name:
            self.manager.wczytaj_z_excela(file_name)
            self.odswiez_widok()

    def action_wyczysc(self):
        reply = QMessageBox.question(self, 'Potwierdzenie', 'Czy na pewno chcesz wyczyścić wszystkie dane?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.manager.czysc_dane()
            self.odswiez_widok()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = BudgetGUI()
    window.show()
    sys.exit(app.exec_())